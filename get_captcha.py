#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import http.client
import json
import base64
import time
import os
import sys
import pandas as pd
from datetime import datetime
import glob

# 添加ddddocr路径 - 使用相对路径
current_dir = os.path.dirname(os.path.abspath(__file__))
ddddocr_path = os.path.join(current_dir, '..', '..', 'rongshu', 'phoneNumber', 'ddddocr-master')
sys.path.append(ddddocr_path)
import ddddocr

# 导入公共API调用模块
from common_api import create_api_client

def get_captcha_and_save():
    """
    调用验证码接口，获取响应头中的UUID和响应体中的base64数据，
    将base64转换为图片保存到temp_captcha目录
    """
    
    print("🔑 开始获取验证码...")
    print("=" * 50)
    
    try:
        # 确保temp_captcha目录存在 - 使用相对路径
        captcha_dir = os.path.join(current_dir, '..', '..', 'rongshu', 'phoneNumber', 'temp_captcha')
        if not os.path.exists(captcha_dir):
            os.makedirs(captcha_dir)
            print(f"📁 创建目录: {captcha_dir}")
        
        # 使用http.client调用验证码接口
        conn = http.client.HTTPSConnection("www.dianhua.cn")
        payload = ''
        headers = {
            'Pragma': 'no-cache'
        }
        
        print("📡 正在调用验证码接口...")
        conn.request("GET", "/api/1.0.0/dianhua/captcha", payload, headers)
        res = conn.getresponse()
        data = res.read()
        
        print(f"📊 响应状态码: {res.status}")
        
        # 从响应头中获取CAPTCHA-UUID
        response_headers = dict(res.getheaders())
        captcha_uuid = response_headers.get('CAPTCHA-UUID') or response_headers.get('captcha-uuid')
        
        print(f"🆔 响应头中的CAPTCHA-UUID: {captcha_uuid}")
        
        if not captcha_uuid:
            print("❌ 未在响应头中找到CAPTCHA-UUID")
            return None, None
        
        # 解析响应体获取base64数据
        response_text = data.decode("utf-8")
        print(f"📋 响应体内容: {response_text}")
        
        try:
            captcha_data = json.loads(response_text)
        except json.JSONDecodeError as e:
            print(f"❌ 响应体不是有效的JSON: {e}")
            return None, None
        
        if captcha_data.get('code') != 0:
            print(f"❌ 验证码API返回错误: {captcha_data}")
            return None, None
        
        # 获取base64数据
        base64_data = captcha_data.get('data')
        if not base64_data:
            print("❌ 响应体中未找到base64数据")
            return None, None
        
        print(f"📸 获取到base64数据长度: {len(base64_data)}")
        
        # 处理base64数据（移除data:image/png;base64,前缀如果存在）
        if base64_data.startswith('data:image/png;base64,'):
            base64_data = base64_data.replace('data:image/png;base64,', '')
            print("🔧 移除了base64前缀")
        
        # 解码base64数据
        try:
            image_data = base64.b64decode(base64_data)
            print(f"✅ base64解码成功，图片数据大小: {len(image_data)} 字节")
        except Exception as e:
            print(f"❌ base64解码失败: {e}")
            return None, None
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        uuid_short = captcha_uuid[:8] if captcha_uuid else "unknown"
        filename = f"captcha_{timestamp}_{uuid_short}.png"
        filepath = os.path.join(captcha_dir, filename)
        
        # 保存图片文件
        try:
            with open(filepath, 'wb') as f:
                f.write(image_data)
            print(f"💾 验证码图片已保存: {filepath}")
        except Exception as e:
            print(f"❌ 保存图片失败: {e}")
            return None, None
        
        print("\n🎉 验证码获取成功!")
        print(f"🆔 CAPTCHA-UUID: {captcha_uuid}")
        print(f"📁 图片路径: {filepath}")
        
        return captcha_uuid, filepath
        
    except Exception as e:
        print(f"❌ 获取验证码过程中发生异常: {e}")
        return None, None
    
    finally:
        try:
            conn.close()
        except:
            pass

def query_phone_number(captcha_uuid, phone_number, captcha_code):
    """
    使用获取到的CAPTCHA-UUID查询电话号码信息
    
    Args:
        captcha_uuid: 从验证码接口响应头获取的UUID
        phone_number: 要查询的手机号
        captcha_code: 用户输入的验证码
    
    Returns:
        dict: 查询结果数据，包含success状态和data信息
    """
    
    print(f"\n🔍 开始查询电话号码信息...")
    print("=" * 50)
    print(f"📱 手机号: {phone_number}")
    print(f"🔑 验证码: {captcha_code}")
    print(f"🆔 CAPTCHA-UUID: {captcha_uuid}")
    
    try:
        # 使用http.client进行查询
        conn = http.client.HTTPSConnection("www.dianhua.cn")
        payload = ''
        headers = {
            'CAPTCHA-UUID': captcha_uuid,
            'Pragma': 'no-cache'
        }
        
        # 构建查询URL
        query_url = f"/api/1.0.0/dianhua/searchTel?tel={phone_number}&code={captcha_code}"
        print(f"📡 查询URL: {query_url}")
        
        conn.request("GET", query_url, payload, headers)
        res = conn.getresponse()
        data = res.read()
        
        print(f"📊 响应状态码: {res.status}")
        
        # 解析响应
        response_text = data.decode("utf-8")
        print(f"📋 响应内容: {response_text}")
        
        try:
            import json
            result_data = json.loads(response_text)
            
            if result_data.get('code') == 0:
                print("🎉 查询成功!")
                data_info = result_data.get('data', {})
                if data_info:
                    print("📱 电话信息:")
                    if isinstance(data_info, dict):
                        for key, value in data_info.items():
                            print(f"   {key}: {value}")
                    else:
                        print(f"   {data_info}")
                else:
                    print("📱 未找到相关信息")
                
                return {
                    'success': True,
                    'data': data_info,
                    'phone_number': phone_number,
                    'message': result_data.get('message', '')
                }
            else:
                error_msg = result_data.get('message', '未知错误')
                print(f"❌ 查询失败: {error_msg}")
                return {
                    'success': False,
                    'data': {},
                    'phone_number': phone_number,
                    'message': error_msg
                }
                
        except json.JSONDecodeError:
            print(f"❌ 响应不是有效的JSON格式")
            return {
                'success': False,
                'data': {},
                'phone_number': phone_number,
                'message': '响应格式错误'
            }
            
    except Exception as e:
        print(f"❌ 查询过程中发生异常: {e}")
        return {
            'success': False,
            'data': {},
            'phone_number': phone_number,
            'message': f'查询异常: {str(e)}'
        }
    
    finally:
        try:
            conn.close()
        except:
            pass

def recognize_captcha_with_ocr(image_path):
    """
    使用ddddocr识别验证码
    
    Args:
        image_path: 验证码图片路径
    
    Returns:
        str: 识别出的验证码文本，如果识别失败返回None
    """
    
    print(f"\n🤖 开始OCR识别验证码...")
    print("=" * 50)
    print(f"📁 图片路径: {image_path}")
    
    try:
        # 初始化ddddocr
        print("🔧 初始化OCR引擎...")
        ocr = ddddocr.DdddOcr(show_ad=False)  # 不显示广告
        
        # 读取图片文件
        with open(image_path, 'rb') as f:
            image_data = f.read()
        
        print(f"📖 读取图片数据，大小: {len(image_data)} 字节")
        
        # 执行OCR识别
        print("🔍 正在识别验证码...")
        result = ocr.classification(image_data)
        
        print(f"✅ OCR识别结果: {result}")
        
        if result and len(result.strip()) > 0:
            return result.strip()
        else:
            print("❌ OCR识别结果为空")
            return None
            
    except Exception as e:
        print(f"❌ OCR识别异常: {e}")
        return None

def get_latest_number_list_file():
    """
    获取最新的numberList JSON文件路径
    
    Returns:
        str: 最新的JSON文件路径，如果没有找到则返回None
    """
    
    # 使用相对路径
    files_dir = "files"
    pattern = os.path.join(files_dir, "numberList_*.json")
    
    try:
        # 查找所有匹配的文件
        json_files = glob.glob(pattern)
        
        if not json_files:
            print(f"❌ 在 {files_dir} 目录中未找到 numberList_*.json 文件")
            return None
        
        # 按修改时间排序，获取最新的文件
        latest_file = max(json_files, key=os.path.getmtime)
        print(f"📁 找到最新的numberList文件: {latest_file}")
        
        return latest_file
        
    except Exception as e:
        print(f"❌ 查找numberList文件时出错: {e}")
        return None

def load_phone_numbers_from_json(json_path):
    """
    从JSON文件中读取手机号码列表
    
    Args:
        json_path: JSON文件路径
    
    Returns:
        list: 手机号码列表
    """
    
    print(f"📊 读取JSON文件: {json_path}")
    
    try:
        # 读取JSON文件
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 从新的JSON格式中提取numberList
        if isinstance(data, dict) and 'numberList' in data:
            phone_numbers = data['numberList']
            print(f"📏 从numberList中读取到 {len(phone_numbers)} 个号码")
        else:
            # 兼容旧格式（直接是数组）
            phone_numbers = data if isinstance(data, list) else []
            print(f"📏 使用兼容模式，读取到 {len(phone_numbers)} 个号码")
        
        # 确保所有号码都是字符串格式，保持原始格式不变
        formatted_phones = []
        for phone in phone_numbers:
            if phone:  # 去除空值
                phone_str = str(phone).strip()
                formatted_phones.append(phone_str)
        
        print(f"📱 成功读取 {len(formatted_phones)} 个手机号码")
        print(f"📋 前3个号码示例: {formatted_phones[:3]}")
        return formatted_phones
            
    except Exception as e:
        print(f"❌ 读取JSON文件失败: {e}")
        return []

def load_phone_numbers_from_excel(excel_path):
    """
    从Excel文件中读取手机号码列表（保留原函数以兼容）
    
    Args:
        excel_path: Excel文件路径
    
    Returns:
        list: 手机号码列表
    """
    
    print(f"📊 读取Excel文件: {excel_path}")
    
    try:
        # 读取Excel文件
        df = pd.read_excel(excel_path)
        
        print(f"📋 Excel文件列名: {df.columns.tolist()}")
        print(f"📏 数据行数: {len(df)}")
        
        # 获取号码列
        if '号码' in df.columns:
            phone_numbers = df['号码'].tolist()
            # 转换为字符串并保留前导0，去除空值
            formatted_phones = []
            for phone in phone_numbers:
                if pd.notna(phone):
                    # 如果是数字类型，需要特殊处理保留前导0
                    if isinstance(phone, (int, float)):
                        # 转换为字符串，如果原始数据有前导0，需要补回来
                        phone_str = str(int(phone))  # 先转为int去除小数点，再转字符串
                        # 检查是否需要补前导0
                        # 如果是10位数字且不以1开头，很可能原来有前导0
                        if len(phone_str) == 10:
                            phone_str = '0' + phone_str
                        # 如果是9位数字，可能原来有两个前导0
                        elif len(phone_str) == 9:
                            phone_str = '00' + phone_str
                        # 如果是8位数字，可能原来有三个前导0  
                        elif len(phone_str) == 8:
                            phone_str = '000' + phone_str
                    else:
                        phone_str = str(phone).strip()
                    
                    formatted_phones.append(phone_str)
            
            print(f"📱 成功读取 {len(formatted_phones)} 个手机号码")
            print(f"📋 前3个号码示例: {formatted_phones[:3]}")
            return formatted_phones
        else:
            print(f"❌ Excel文件中未找到'号码'列")
            return []
            
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        return []

def save_single_result_to_excel(result, output_path):
    """
    将单条查询结果追加保存到Excel文件
    
    Args:
        result: 单条查询结果
        output_path: 输出Excel文件路径
    """
    
    print(f"\n💾 追加单条结果到Excel文件: {output_path}")
    
    try:
        from openpyxl import load_workbook, Workbook
        
        # 准备单条数据，确保手机号为字符串类型
        if result['success'] and result['data']:
            data_info = result['data']
            row_data = [
                str(result['phone_number']),  # 手机号
                '成功',  # 查询状态
                data_info.get('telnum', ''),  # telnum
                data_info.get('name', ''),  # name
                data_info.get('flag', ''),  # flag
                data_info.get('id', ''),  # id
                data_info.get('teltype', ''),  # teltype
                ''  # 错误信息
            ]
        else:
            row_data = [
                str(result['phone_number']),  # 手机号
                '失败',  # 查询状态
                '',  # telnum
                '',  # name
                '',  # flag
                '',  # id
                '',  # teltype
                result['message']  # 错误信息
            ]
        
        # 检查文件是否存在
        if os.path.exists(output_path):
            # 文件存在，加载现有工作簿
            wb = load_workbook(output_path)
            ws = wb.active
        else:
            # 文件不存在，创建新工作簿
            wb = Workbook()
            ws = wb.active
            # 添加表头
            headers = ['手机号', '查询状态', 'telnum', 'name', 'flag', 'id', 'teltype', '错误信息']
            ws.append(headers)
        
        # 添加数据行
        ws.append(row_data)
        
        # 设置手机号列（第1列）为文本格式
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.number_format = '@'  # 文本格式
        
        # 保存文件
        wb.save(output_path)
        
        print(f"✅ 成功追加1条结果到 {output_path}")
        print(f"📱 手机号: {result['phone_number']}, 状态: {'成功' if result['success'] else '失败'}")
        
    except Exception as e:
        print(f"❌ 追加保存Excel文件失败: {e}")

def initialize_excel_file(output_path):
    """
    初始化Excel文件，创建表头
    
    Args:
        output_path: 输出Excel文件路径
    """
    
    print(f"\n📋 初始化Excel文件: {output_path}")
    
    try:
        from openpyxl import Workbook
        
        # 如果文件已存在，不需要重新初始化
        if os.path.exists(output_path):
            print(f"📄 Excel文件已存在，跳过初始化")
            return True
        
        # 创建新工作簿
        wb = Workbook()
        ws = wb.active
        
        # 添加表头
        headers = ['手机号', '查询状态', 'telnum', 'name', 'flag', 'id', 'teltype', '错误信息']
        ws.append(headers)
        
        # 设置手机号列（第1列）的标题行为文本格式
        cell = ws.cell(row=1, column=1)
        cell.number_format = '@'
        
        # 保存文件
        wb.save(output_path)
        
        print(f"✅ 成功初始化Excel文件")
        return True
        
    except Exception as e:
        print(f"❌ 初始化Excel文件失败: {e}")
        return False

def save_results_to_excel(results, output_path):
    """
    将查询结果保存到Excel文件（批量模式，保持向后兼容）
    
    Args:
        results: 查询结果列表
        output_path: 输出Excel文件路径
    """
    
    print(f"\n💾 保存结果到Excel文件: {output_path}")
    
    try:
        # 准备数据
        data_rows = []
        
        for result in results:
            if result['success'] and result['data']:
                data_info = result['data']
                row = {
                    '手机号': result['phone_number'],
                    '查询状态': '成功',
                    'telnum': data_info.get('telnum', ''),
                    'name': data_info.get('name', ''),
                    'flag': data_info.get('flag', ''),
                    'id': data_info.get('id', ''),
                    'teltype': data_info.get('teltype', ''),
                    '错误信息': ''
                }
            else:
                row = {
                    '手机号': result['phone_number'],
                    '查询状态': '失败',
                    'telnum': '',
                    'name': '',
                    'flag': '',
                    'id': '',
                    'teltype': '',
                    '错误信息': result['message']
                }
            
            data_rows.append(row)
        
        # 创建DataFrame并保存
        df = pd.DataFrame(data_rows)
        df.to_excel(output_path, index=False)
        
        print(f"✅ 成功保存 {len(data_rows)} 条结果到 {output_path}")
        
        # 显示统计信息
        success_count = sum(1 for r in results if r['success'])
        fail_count = len(results) - success_count
        print(f"📊 查询统计: 成功 {success_count} 条，失败 {fail_count} 条")
        
    except Exception as e:
        print(f"❌ 保存Excel文件失败: {e}")

def batch_query_phones(phone_numbers, output_path, max_retries=5):
    """
    批量查询手机号码信息，每成功一条就立即保存到Excel
    
    Args:
        phone_numbers: 手机号码列表
        output_path: 输出Excel文件路径
        max_retries: 最大重试次数
    
    Returns:
        list: 查询结果列表
    """
    
    print(f"\n🚀 开始批量查询 {len(phone_numbers)} 个手机号码")
    print("=" * 70)
    
    # 初始化Excel文件
    if not initialize_excel_file(output_path):
        print("❌ 初始化Excel文件失败，程序退出")
        return []
    
    results = []
    
    for i, phone_number in enumerate(phone_numbers, 1):
        print(f"\n📋 处理第 {i}/{len(phone_numbers)} 个号码: {phone_number}")
        print("-" * 50)
        
        retry_count = 0
        success = False
        used_images = []  # 记录本次查询使用的验证码图片
        
        while retry_count < max_retries and not success:
            try:
                # 步骤1: 获取验证码
                print(f"🔄 第 {retry_count + 1} 次尝试")
                uuid, image_path = get_captcha_and_save()
                
                if not uuid or not image_path:
                    print(f"❌ 获取验证码失败")
                    retry_count += 1
                    continue
                
                # 记录使用的图片路径
                used_images.append(image_path)
                
                # 步骤2: OCR识别验证码
                captcha_code = recognize_captcha_with_ocr(image_path)
                
                if not captcha_code:
                    print(f"❌ OCR识别失败")
                    retry_count += 1
                    continue
                
                # 步骤3: 查询手机号
                result = query_phone_number(uuid, phone_number, captcha_code)
                
                if result['success']:
                    print(f"✅ 查询成功!")
                    results.append(result)
                    
                    # 立即保存成功的结果到Excel
                    save_single_result_to_excel(result, output_path)
                    
                    # 检查是否有flag值，如果有则调用公共API
                    if result.get('data') and result['data'].get('flag'):
                        flag_value = result['data']['flag']
                        print(f"🔍 检测到flag值: {flag_value}")
                        
                        try:
                            # 创建API客户端并调用
                            api_client = create_api_client()
                            tag = f"号码邦-{flag_value}"
                            
                            print(f"📞 调用公共API...")
                            print(f"   📱 Number: {phone_number}")
                            print(f"   🏷️  Tag: {tag}")
                            
                            api_result = api_client.call_api_with_number_tag(phone_number, tag)
                            
                            if api_result.get('success'):
                                print(f"✅ 公共API调用成功!")
                            else:
                                print(f"❌ 公共API调用失败: {api_result.get('error', '未知错误')}")
                                
                        except Exception as api_e:
                            print(f"❌ 调用公共API时发生异常: {api_e}")
                    else:
                        print(f"ℹ️  未检测到flag值，跳过公共API调用")
                    
                    success = True
                else:
                    print(f"❌ 查询失败: {result['message']}")
                    if '验证码' in result['message'] or 'captcha' in result['message'].lower():
                        # 验证码错误，重新获取验证码重试
                        print(f"🔄 验证码错误，将重新获取验证码并重试")
                        retry_count += 1
                    else:
                        # 其他错误，不重试，但也要保存失败结果
                        results.append(result)
                        save_single_result_to_excel(result, output_path)
                        success = True
                
                # 只有成功时才添加延迟，失败重试时不延迟
                if success and i < len(phone_numbers):  # 不是最后一个
                    print("⏱️  等待2秒...")
                    time.sleep(2)
                elif not success and retry_count < max_retries:
                    # 重试前短暂延迟
                    print("⏱️  重试前等待1秒...")
                    time.sleep(1)
                    
            except Exception as e:
                print(f"❌ 处理异常: {e}")
                retry_count += 1
        
        # 清理本次查询使用的验证码图片
        for img_path in used_images:
            try:
                if os.path.exists(img_path):
                    os.remove(img_path)
                    print(f"🗑️  已删除验证码图片: {os.path.basename(img_path)}")
            except Exception as e:
                print(f"⚠️  删除图片失败 {img_path}: {e}")
        
        # 如果所有重试都失败了
        if not success:
            error_result = {
                'success': False,
                'data': {},
                'phone_number': phone_number,
                'message': f'重试{max_retries}次后仍然失败'
            }
            results.append(error_result)
            # 保存失败结果到Excel
            save_single_result_to_excel(error_result, output_path)
    
    return results

def main():
    """主函数 - 批量处理模式"""
    print("🚀 批量验证码自动识别和查询工具")
    print("功能：1. 从JSON读取手机号  2. 批量OCR识别  3. 批量查询并实时保存结果")
    print("=" * 70)
    
    # 获取最新的numberList JSON文件
    input_json_path = get_latest_number_list_file()
    if not input_json_path:
        print("❌ 未找到有效的numberList JSON文件，程序退出")
        return False
    
    # 输出Excel路径（相对路径）
    output_excel_path = "手机号查询结果.xlsx"
    
    # 步骤1: 从JSON读取手机号码
    print(f"\n📋 步骤1: 从JSON文件读取手机号码")
    phone_numbers = load_phone_numbers_from_json(input_json_path)
    
    if not phone_numbers:
        print(f"\n❌ 未能读取到手机号码，程序退出")
        return False
    
    print(f"\n✅ 成功读取 {len(phone_numbers)} 个手机号码")
    print(f"📱 前5个号码预览: {phone_numbers[:5]}")
    
    # 步骤2: 批量查询（每成功一条就保存）
    print(f"\n📋 步骤2: 开始批量查询（实时保存模式）")
    results = batch_query_phones(phone_numbers, output_excel_path)
    
    if not results:
        print(f"\n❌ 批量查询失败，程序退出")
        return False
    
    # 显示最终统计信息
    print(f"\n🎉 批量查询完成!")
    print(f"📊 输入文件: {input_json_path}")
    print(f"📊 输出文件: {output_excel_path}")
    
    # 显示统计信息
    success_count = sum(1 for r in results if r['success'])
    fail_count = len(results) - success_count
    print(f"📊 最终统计: 成功 {success_count} 条，失败 {fail_count} 条")
    
    return True

if __name__ == "__main__":
    main()